#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
exports for dessia_common

"""
from copy import copy
from string import Template
from typing import List, Any, Union


from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook
import openpyxl.utils
from dessia_common.breakdown import breakdown


def is_hashable(value):
    """Determine whether `value` can be hashed."""
    try:
        hash(value)
    except TypeError:
        return False
    return True


def is_number(value):
    """
    Determine if the value is a int or a float
    """
    return isinstance(value, (int, float))


def is_builtins_list(list_):
    """
    Determin if a list is only composed of builtins
    """
    for element in list_:
        if not (is_number(element) or isinstance(element, str)):
            return False
    return True


class XLSXWriter:

    max_column_width = 40
    color_dessIA1 = "263238"
    color_dessIA2 = "537CB0"
    grey1 = "f1f1f1"
    white_font = Font(color="FFFFFF")

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    def __init__(self, object_):
        """
        :param object_: an Dessiaobject to write as excel
        """

        self.pattern_color1 = PatternFill(
            fill_type="solid",
            start_color=self.color_dessIA1,
            end_color=self.color_dessIA1)

        self.pattern_color2 = PatternFill(
            fill_type="solid",
            start_color=self.color_dessIA2,
            end_color=self.color_dessIA2)

        self.workbook = Workbook()
        self.main_sheet = self.workbook.active
        self.object = object_

        self.paths = breakdown(object_)

        self.classes_to_sheets = {}
        self.object_to_sheet_row = {}
        for class_name, obj_paths in self.paths.items():
            sheet = self.workbook.create_sheet(class_name)
            self.classes_to_sheets[class_name] = sheet

            for i, (obj, path) in enumerate(obj_paths.items()):
                self.object_to_sheet_row[obj] = (sheet, i + 2, path)

        self.write()

    def write_class_header_to_row(self, obj_of_class, sheet, row_number):
        """
        Writes to a sheet the class header: finds columns names from a class
        """
        cell = sheet.cell(row=row_number, column=1, value='Path')
        cell.fill = self.pattern_color2
        cell.border = self.thin_border
        cell.font = self.white_font

        cell = sheet.cell(row=row_number, column=2, value='name')
        cell.fill = self.pattern_color2
        cell.border = self.thin_border
        cell.font = self.white_font
        i = 3

        for (k, _) in sorted(obj_of_class.__dict__.items()):
            if (not k.startswith('_')) and k != 'name':
                cell = sheet.cell(row=row_number, column=i, value=str(k))

                cell.border = self.thin_border
                cell.fill = self.pattern_color2
                cell.font = self.white_font
                i += 1

    def write_value_to_cell(self, value, sheet, row_number, column_number):
        """
        Write a given value to a cell. Insert it as a link if it is an object
        """
        cell_link = None
        if isinstance(value, dict):
            str_v = f'Dict of {len(value)} items'
        elif isinstance(value, list):
            if is_builtins_list(value):
                str_v = str(value)
            else:
                str_v = f'List of {len(value)} items'

        elif isinstance(value, set):
            str_v = f'Set of {len(value)} items'
        elif isinstance(value, float):
            str_v = round(value, 6)
        elif is_hashable(value) and value in self.object_to_sheet_row:
            ref_sheet, ref_row_number, ref_path = self.object_to_sheet_row[value]
            str_v = ref_path
            cell_link = f'#{ref_sheet.title}!A{ref_row_number}'
        else:
            str_v = str(value)

        cell = sheet.cell(row=row_number, column=column_number, value=str_v)
        if cell_link:
            cell.hyperlink = cell_link

        cell.border = self.thin_border

    def write_object_to_row(self, obj, sheet, row_number, path=''):
        """
        Write on object to a row. Loops on its attributes to write its value in each cell
        """
        cell = sheet.cell(row=row_number, column=1, value=path)
        cell.border = self.thin_border
        if hasattr(obj, 'name'):
            cell = sheet.cell(row=row_number, column=2, value=obj.name)
        else:
            cell = sheet.cell(row=row_number, column=2, value='No name in model')

        cell.border = self.thin_border
        i = 3
        for (k, value) in sorted(obj.__dict__.items()):
            if (not k.startswith('_')) and k != 'name':
                self.write_value_to_cell(value, sheet, row_number, i)

                i += 1

                # column_width = min((len(k) + 1.5), max_column_width)
                # column_name = openpyxl.utils.cell.get_column_letter(i)
                # sheet.column_dimensions[column_name].width = column_width

    def write_object_id(self, sheet):
        """
        Write object id to a given sheet
        """
        sheet.title = f'Object {self.object.__class__.__name__}'

        sheet['A1'] = 'Module'
        sheet['B1'] = 'Class'
        sheet['C1'] = 'name'

        sheet['A1'].border = self.thin_border
        sheet['B1'].border = self.thin_border
        sheet['C1'].border = self.thin_border
        sheet['A1'].fill = self.pattern_color1
        sheet['B1'].fill = self.pattern_color1
        sheet['C1'].fill = self.pattern_color1
        sheet['A1'].font = self.white_font
        sheet['B1'].font = self.white_font
        sheet['C1'].font = self.white_font

        sheet['A2'] = self.object.__module__
        sheet['B2'] = self.object.__class__.__name__
        sheet['C2'] = self.object.name

        sheet['A2'].border = self.thin_border
        sheet['B2'].border = self.thin_border
        sheet['C2'].border = self.thin_border
        sheet['A2'].fill = self.pattern_color1
        sheet['B2'].fill = self.pattern_color1
        sheet['C2'].fill = self.pattern_color1
        sheet['A2'].font = self.white_font
        sheet['B2'].font = self.white_font
        sheet['C2'].font = self.white_font

        sheet['A3'] = 'Attribute'
        sheet['A4'] = 'Value'
        sheet['A3'].border = self.thin_border
        sheet['A4'].border = self.thin_border

    def write(self):
        """
        Generate the whole file
        """
        # name_column_width = 0
        self.write_object_id(self.main_sheet)
        self.write_class_header_to_row(self.object, self.main_sheet, 3)
        self.write_object_to_row(self.object, self.main_sheet, 4)
        self.autosize_sheet_columns(self.main_sheet, 5, 30)

        for class_name, obj_paths in self.paths.items():
            sheet = self.classes_to_sheets[class_name]

            for obj, path in obj_paths.items():
                _, row_number, path = self.object_to_sheet_row[obj]
                self.write_object_to_row(obj, sheet, row_number, path)
            self.write_class_header_to_row(obj, sheet, 1)

            sheet.auto_filter.ref = f"A1:{openpyxl.utils.cell.get_column_letter(sheet.max_column)}{len(obj_paths) + 1}"
            self.autosize_sheet_columns(sheet, 5, 30)

    def save_to_file(self, filepath: str):
        """
        Save to a filepath (open) and write
        """
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'
            print(f"Changing name to {filepath}")

        with open(filepath, 'rb') as file:
            self.save_to_stream(file)

    def save_to_stream(self, stream):
        """
        Saves the file to a binary stream
        """
        self.workbook.save(stream)

    @staticmethod
    def autosize_sheet_columns(sheet, min_width=5, max_width=30):
        """
        Autosize the sheet columns by analyzing the content. Min and max width must be specified
        """
        # Autosize columns
        for col in sheet.columns:
            width = min_width
            column = col[1].column_letter  # Get the column name
            # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based)
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > width:
                        width = len(str(cell.value))
                except AttributeError:
                    pass
            if width > 0:
                adjusted_width = min((width + 0.5), max_width)
                sheet.column_dimensions[column].width = adjusted_width


class MarkdownWriter:
    def __init__(self, object_):
        self.object_ = object_
        # self.text = self.to_text()

    def _class_summary(self):
        return ("Summary: This is a standard class summary and can be customized by changing method " +
                "<_class_summary()> of DessiaObject to write a class summary in markdown.\n" +
                "More information can be found here: https://www.markdownguide.org/cheat-sheet/")

    def _titles(self):
        return "| Attribute | Type | Contains | Subvalues |\n"

    def _empty_row(self):
        return "| ------ | ------ | ------ | ------ |\n"

    def _sequence_row(self, value):
        if len(value) == 0:
            return {}, copy(value)

        if isinstance(value, dict):
            in_values = list(value.values())
        else:
            in_values = value

        all_class = set(subvalue.__class__.__name__ for subvalue in in_values)

        return all_class

    def _printed_string_in_table(self, printed_string, str_types):
        printed_string = printed_string[:20] + ('...' if len(printed_string) > 20 else '')
        return str_types + f" {printed_string} |\n"

    def _simple_value_row(self, value):
        if hasattr(value, 'name'):
            printed_string = (value.name if value.name != '' else 'unnamed')
        else:
            printed_string = str(value)
        return self._printed_string_in_table(printed_string, ' - |')

    def _multiclass_row(self, value, all_class):
        if hasattr(value, 'name'):
            printed_string = [subvalue.name if subvalue.name != '' else 'unnamed' for subvalue in value]
            printed_string = ', '.join(printed_string)
        else:
            printed_string = str(value)

        str_all_class = str(all_class).translate(str(all_class).maketrans('', '', "{}'"))
        str_types = f" {len(value)} elements of classes {str_all_class} |"
        return self._printed_string_in_table(printed_string, str_types)

    def _attr_table(self):
        table_attributes = self._titles()
        table_attributes += self._empty_row()
        for attr, value in self.object_.__dict__.items():
            table_attributes += f"| {attr} | {value.__class__.__name__} |"
            all_class = {}

            if isinstance(value, (list, tuple, dict)):
                all_class = self._sequence_row(value)

            if len(all_class) == 0:
                table_attributes += self._simple_value_row(value)
            else:
                table_attributes += self._multiclass_row(value, all_class)

        return table_attributes

    def to_text(self) -> str:
        """
        Render a markdown of the object output type: string
        """
        printed_name = (self.object_.name + ' ' if self.object_.name != '' else '')
        text = f"# Object {printed_name}of class {self.__class__.__name__}\n\n"
        text += "## Summary\n"
        text += "\n$summary\n\n"
        text += "\n## Attribute values\n\n"
        text += "$table_attributes\n"
        text = Template(text).substitute(summary=self._class_summary(),
                                         table_attributes=self._attr_table(),
                                         name=self.object_.name, class_=self.object_.__class__.__name__)
        return text

    def write_table(self, object_):
        if isinstance(object_, (float, int, bool, complex, str)):
            return self._simple_value_table(object_)
        if isinstance(object_, (list, dict, set)):
            return self._sequence_table(object_)
        if hasattr(object_, 'name'):
            return self._object_with_name_table(object_)
        else:
            return self._object_table(object_)

    def _write_head_table(self, col_names: List[str] = None) -> str:
        return ("| " + " | ".join(col_names) + " |\n" +
                "| ------ " * len(col_names) + "|\n")

    def _sequence_to_str(self, value: List[Union[list, dict, set]]):
        in_values = value
        str_types = f" {len(value)} elements"
        if isinstance(value, dict):
            in_values = list(value.values())

        all_class = set(subvalue.__class__.__name__ for subvalue in in_values)

        if len(all_class) == 1:
            str_types += f"of classes {all_class[0].__name__} |"
            if hasattr(all_class[0], 'name'):
                printed_string = [subvalue.name if subvalue.name != '' else 'unnamed' for subvalue in value]
                printed_string = ', '.join(printed_string)
            else:
                printed_string = str(value)

        else:
            str_all_class = str(all_class).translate(str(all_class).maketrans('', '', "{}'"))
            str_types += f"of classes {str_all_class} |"

        return self._printed_string_in_table(printed_string, str_types)


    def _write_line_table(self, row: List[Any]) -> str:
        line = "|"
        for value in row:
            line += " "
            if isinstance(value, (float, int, bool, complex)):
                line += str(round(value, 6))

            if isinstance(value, str):
                line += value[:15]

            if isinstance(value, (list, dict, set)):
                line += self._sequence_to_str(value)

            if hasattr(value, 'name'):
                return self._object_with_name_line(value)

            # else:
            #     return self._object_line(value)

            line += " |"

        return line + "\n"

    def _write_content_table(self, content: List[List[Any]]) -> str:
        table = ''
        for row in content[:5]:
            table += self._write_line_table(row)
        return table


    def _matrix_table(self, matrix: List[List[float]], col_names: List[str] = None) -> str:
        table = self._write_head_table(col_names)
        table += self._write_content_table(matrix)
        return table

    def write_table(self, content, head):
        return self._matrix_table(content, head)

